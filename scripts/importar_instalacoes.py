"""
Script para importar dados da planilha "INSTALAÇÕES GERAIS DE TECNICOS" para o Supabase.
Normaliza a ordem das colunas (que muda entre abas), datas, e insere em lotes.
"""
from __future__ import annotations
import openpyxl
import json
import re
from datetime import datetime
import urllib.request

SUPABASE_URL = "https://jlrslrljvpveaeheetlm.supabase.co"
SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImpscnNscmxqdnB2ZWFlaGVldGxtIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NDAzNDM2NSwiZXhwIjoyMDg5NjEwMzY1fQ._iwyZf5vBiMeeh_9wg3SxCT5UEWHsXBIo42xogJpTeg"

SPREADSHEET_PATH = "/Users/matheussouza/Downloads/castelo das ilusões/INSTALAÇÕES GERAIS DE TECNICOS .xlsx"

# Abas para ignorar
SKIP_SHEETS = ["Respostas ao formulário 2"]

# Mapeamento de nomes de cabeçalho para campos normalizados
HEADER_MAP = {
    "DATA": "data",
    "DATA DE INSTALAÇÃO": "data",
    "TECNICO": "tecnico_nome",
    "TÉCNICO": "tecnico_nome",
    "PLACA": "placa",
    "PLACA DO VEÍCULO": "placa",
    "FILIAL": "filial",
    "FILIAL/LOCAL": "filial",
    "VALOR": "valor",
    "OBSERVAÇÃO": "observacao",
    "OBSERVAÇÕES": "observacao",
}


def normalize_header(header: str) -> str | None:
    """Map a spreadsheet header to its normalized field name."""
    if header is None:
        return None
    clean = header.strip().upper()
    # Remove trailing colons/spaces
    clean = re.sub(r'[\s:]+$', '', clean)
    return HEADER_MAP.get(clean)


def parse_date(val) -> str | None:
    """Convert various date formats to YYYY-MM-DD."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # Try datetime string like "2025-10-01 00:00:00"
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        pass
    # Try "dd/mm/yyyy" with optional suffix like "(SABADO)"
    match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if match:
        d, m, y = match.groups()
        try:
            return datetime(int(y), int(m), int(d)).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def parse_valor(val) -> float:
    """Parse a monetary value."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    try:
        return float(s)
    except ValueError:
        return 0


def clean_text(val) -> str:
    """Clean a text value."""
    if val is None:
        return ""
    return str(val).strip()


def supabase_insert_batch(records: list[dict], batch_size=50):
    """Insert records into Supabase in batches using REST API."""
    url = f"{SUPABASE_URL}/rest/v1/instalacoes"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    total = len(records)
    inserted = 0
    errors = 0

    for i in range(0, total, batch_size):
        batch = records[i:i + batch_size]
        data = json.dumps(batch).encode('utf-8')
        req = urllib.request.Request(url, data=data, headers=headers, method='POST')
        try:
            with urllib.request.urlopen(req) as resp:
                inserted += len(batch)
                print(f"  Inseridos {inserted}/{total} registros...")
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            print(f"  ERRO no lote {i}-{i+len(batch)}: {e.code} - {body}")
            errors += len(batch)

    return inserted, errors


def main():
    print(f"Abrindo planilha: {SPREADSHEET_PATH}")
    wb = openpyxl.load_workbook(SPREADSHEET_PATH, data_only=True)

    all_records = []
    counter = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"\n[SKIP] {sheet_name}")
            continue

        ws = wb[sheet_name]
        print(f"\n[PROCESSANDO] {sheet_name}")

        # Map column indices to field names
        col_map = {}
        for col_idx, cell in enumerate(ws[1]):
            field = normalize_header(cell.value if cell.value else None)
            if field:
                col_map[col_idx] = field

        if not col_map:
            print(f"  Sem colunas mapeadas, pulando...")
            continue

        print(f"  Mapeamento: {col_map}")

        sheet_count = 0
        for row in ws.iter_rows(min_row=2, max_col=max(col_map.keys()) + 1):
            raw = {}
            for col_idx, field in col_map.items():
                if col_idx < len(row):
                    raw[field] = row[col_idx].value

            # Skip empty rows
            if not any(raw.get(f) for f in ["data", "placa", "tecnico_nome"]):
                continue

            # Parse and normalize
            data_parsed = parse_date(raw.get("data"))
            placa = clean_text(raw.get("placa"))
            tecnico = clean_text(raw.get("tecnico_nome"))
            filial = clean_text(raw.get("filial"))
            valor = parse_valor(raw.get("valor"))
            observacao = clean_text(raw.get("observacao"))

            # Skip rows without minimum data
            if not placa and not tecnico:
                continue

            counter += 1
            record = {
                "codigo": f"IMP-{counter:05d}",
                "data": data_parsed,
                "placa": placa,
                "tecnico_nome": tecnico,
                "filial": filial,
                "valor": valor,
                "observacao": observacao,
                "imei": "",
                "chip": "",
                "status": "concluida",
                "localizacao_confirmacao": "",
            }
            all_records.append(record)
            sheet_count += 1

        print(f"  {sheet_count} registros extraidos")

    print(f"\n{'='*60}")
    print(f"Total de registros para importar: {len(all_records)}")
    print(f"{'='*60}")

    if not all_records:
        print("Nenhum registro para importar!")
        return

    # Show sample
    print("\nAmostra (primeiros 3 registros):")
    for r in all_records[:3]:
        print(f"  {r}")

    print(f"\nInserindo no Supabase...")
    inserted, errors = supabase_insert_batch(all_records, batch_size=100)
    print(f"\n{'='*60}")
    print(f"RESULTADO: {inserted} inseridos, {errors} erros")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
